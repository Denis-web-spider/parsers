from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import requests
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

start = time.time()

class Parser:
    '''
    Парсер для сайта по продаже автомобильных запчастей https://baza.drom.ru/user/Evomotus/sell_spare_parts/+/%C0%CA%CF%CF.
    Создает в рабочем каталоге папку "img", в которой находятся папки с фотографиями(если они есть) каждого автомобиля.
    Извлекает всю информацию про каждый автомобить и помещает всю информацию в таблицу exel, котороя создается в рабочей директории "Info.xlsx".
    '''

    def __init__(self):
        '''
        Запускает браузер Googl в "безголовом" режиме.
        Созает папку в рабочей директории, котороя имеет имя "img".
        '''

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        google_exe = 'chromedriver.exe'

        self.info = {}

        self.domen = 'https://baza.drom.ru'

        self.driver = webdriver.Chrome(google_exe, options=options)

        current_path = os.getcwd()

        self.imgs_folder = 'imgs'
        path = os.path.join(current_path, '\{}'.format(self.imgs_folder))
        if not os.path.exists(path):
            os.makedirs(path)

    def pages(self, max_page=930):
        '''
        Переходит по страницам сайта и передает html страницы.
        Извлекает url каждого товара.
        '''

        for i in range(1, max_page + 1):
            time.sleep(2)
            url = 'https://baza.drom.ru/user/Evomotus/sell_spare_parts/+/%C0%CA%CF%CF/?page={}'.format(i)
            self.driver.get(url)
            html = self.driver.page_source
            self.gather_urls(html)

    def gather_urls(self, html):
        '''
        Извлекает url каждого товара.
        '''

        soup = BeautifulSoup(html, 'lxml')
        for i in soup.find_all('a', {'class':'bulletinLink bull-item__self-link auto-shy'}):
            self.info[self.domen + i['href']] = {}
            self.info[self.domen + i['href']]['name'] = i.get_text()

    def get_info(self, url):
        '''
        Извлекает всю информацию про деталь автомобиль и помещает её в "self.info[url][характеристика] = значение".
        Извлекает все ссылки на фотографии и помещает их в "self.info[url]['imgs'] = [ссылки]"
        '''

        self.driver.get(url)
        try:
            time.sleep(2)
            button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.LINK_TEXT, 'Показать контакты'))
            )
            action = ActionChains(self.driver)
            action.move_to_element(button)
            action.click()
            action.perform()
        except:
            print("button doesn't exist")

        time.sleep(3)
        soup = BeautifulSoup(self.driver.page_source, 'lxml')
        self.info[url]['цена'] = soup.find('span', {'class': 'viewbull-summary-price__value inplace'}).get_text()

        for i in soup.find_all('div', {'class': 'field viewbull-field__container'}):
            self.info[url][i.find('div', {'class': 'label'}).get_text().strip()] = i.find('div', {'class': 'value'}).get_text().strip()

        for i in soup.find_all('div', {'class': 'field autoPartsModel s-exact-compatibility viewbull-field__container'}):
            self.info[url][i.find('div', {'class': 'label'}).get_text().strip()] = i.find('div', {'class': 'value'}).get_text().strip()

        self.info[url]['Дополнительная информация'] = soup.find('div', {'class': 'bulletinText viewbull-field__container auto-shy'}).get_text().strip()

        for i in soup.find_all('div', {'class': 'field tixt viewbull-field__container'}):
            self.info[url][i.h3.get_text().strip()] = i.div.get_text().strip()

        for i in soup.find_all('div', {'class': 'field text'}):
            self.info[url][i.h3.get_text().strip()] = i.div.get_text().strip()

        contasts = []
        for i in soup.find_all('div', {'class': 'new-contact new-contact_phone'}):
            contasts.append(i.find('div', {'class': 'new-contacts__td new-contact__phone'}).get_text().strip())
        self.info[url]['контакты'] = contasts

        try:
            self.info[url]['email'] = soup.find('div', {'class': 'new-contact new-contact_email'}).get_text().strip()
        except:
            pass

        try:
            imgs = []
            for i in soup.find('div', {'class': 'imagesExSmall'}).find_all('img'):
                imgs.append(i['src'])
            self.info[url]['imgs'] = imgs
        except:
            pass

    def save_imgs(self):
        '''
        Сохраняет фотографии на каждую деталь атомобиля в отдельную папку.
        '''

        for index, url in enumerate(self.info):

            id = index + 1

            new_folder = os.path.join(self.imgs_folder, 'id_{}'.format(id) + '_' + self.info[url]['name'].replace(',', '').replace(' ', '_'))
            if not os.path.exists(new_folder):
                os.makedirs(new_folder)
            try:
                for index, link in enumerate(self.info[url]['imgs']):
                    time.sleep(1)
                    id = index + 1
                    img = requests.get(link, stream=True)
                    img.raw.decode_content = True
                    with open(r'{0}\img_{1}.jpg'.format(new_folder, id), 'wb') as image_file:
                        shutil.copyfileobj(img.raw, image_file)
                    del img
            except:
                pass

    def save_info_in_exel(self):
        '''
        Создает файл exel в рабочей директории и загружает туда всю информацию.
        '''

        wb = Workbook()
        exel = wb.active
        exel.title = 'Info'

        row = 1
        col = 1
        for url in self.info:
            for header in self.info[url]:
                if header != 'imgs':
                    exel.cell(row=row, column=col, value=header).alignment = Alignment(horizontal='center')
                    exel.cell(row=row, column=col).font = Font(size='14')
                    col += 1
            row += 1
            col = 1
            break

        for url in self.info:

            for header in self.info[url]:

                if header == 'контакты':
                    value = '\n'.join(self.info[url][header])
                    exel.cell(row=row, column=col, value=value).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    exel.cell(row=row, column=col).font = Font(size='14')
                    col += 1

                elif header != 'imgs':
                    value = self.info[url][header]
                    exel.cell(row=row, column=col, value=value).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    exel.cell(row=row, column=col).font = Font(size='14')
                    col += 1

            row += 1
            col = 1

        for column_cells in exel.columns:
            length = max(max(map(len, str(cell.value).splitlines())) for cell in column_cells) + 8
            exel.column_dimensions[column_cells[0].column_letter].width = length

        wb.save('Info.xlsx')

    def finish(self):
        '''
        Закрывает браузер.
        '''

        self.driver.quit()

    def execute(self, pages):
        '''
        Выполняет парсинг.
        '''

        try:
            self.pages(pages)

            for i in self.info:
                self.get_info(i)
        except:
            pass

        self.save_info_in_exel()
        self.save_imgs()


parser = Parser()
parser.execute()
