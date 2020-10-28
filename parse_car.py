from selenium import webdriver
from bs4 import BeautifulSoup
import requests
import time
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment

class Parser:
    '''
    Парсер для сайта по продаже автомобилей https://vladivostok.drom.ru/auto/all/?inomarka=1&unsold=1&distance=500&order=price&tcb=1603558101.
    Создает в рабочем каталоге папку "car_img", в которой находятся папки с фотографиями(если они есть) каждого автомобиля.
    Извлекает всю информацию про каждый автомобить и помещает всю информацию в таблицу exel, котороя создается в рабочей директории "Info.xlsx".
    '''

    def __init__(self, main_url):
        '''
        Запускает браузер Googl в "безголовом" режиме и отправляет get запрос на сайт по ссылке "main_url".
        Созает папку в рабочей директории, котороя имеет имя "car_img".
        '''

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        googlexe = 'chromedriver.exe'

        # хеш-таблица для хранения информации об автомобиле
        self.info = {}

        self.driver = webdriver.Chrome(googlexe, options=options)
        self.driver.get(main_url)

        current_path = os.getcwd()

        self.path_name = 'car_img'
        path = os.path.join(current_path, f'\{self.path_name}')
        if not os.path.exists(path):
            os.makedirs(path)

        # переменная предназначенная для индексации названия папок для фотографий машин, он нужен так
        # как попадаются одинаковые модели авто и только модели не достаточно для определения конкретной папки для конкретного авто
        self.folder_id = 1

    def get_car_name_and_url(self):
        '''
        Извлекает ссылки и модель для каждого автомобиля на страничке.
        Ссылки на каждый автомобиль используются в качестве ключа в "info".
        '''

        time.sleep(3)
        self.soup = BeautifulSoup(self.driver.page_source, 'lxml')
        for i in self.soup.find_all('a',{'class':'css-1hgk7d1 eiweh7o2'}):
            self.info[i['href']] = {}
            self.info[i['href']]['name'] = i.find('span', {'data-ftid':'bull_title'}).string

    def next_page(self, numb):
        '''
        Переходит на следующую страницу сайта, в качестве аргумента принимает номер целевой страницы.
        '''

        try:
            self.driver.get('https://vladivostok.drom.ru/auto/all/page{}/?inomarka=1&unsold=1&distance=500&order=price'.format(numb))
        except:
            pass

    def get_car_info(self, car_url):
        '''
        Извлекает всю информацию про автомобиль и помещает её в "self.info[car_url][характеристика] = значение".
        Извлекает все ссылки на фотографии и помещает их в "self.info[car_url]['imgs'] = [ссылки]"
        '''

        self.driver.get(car_url)
        time.sleep(2)
        self.soup = BeautifulSoup(self.driver.page_source, 'lxml')

        self.info[car_url]['Цена'] = self.soup.find('div',{'class':'css-1hu13v1 e162wx9x0'}).contents[0].replace('\\xa', '')

        for i in self.soup.find_all('tr',{'class':'css-10191hq ezjvm5n2'}):
            try:
                self.info[car_url][i.find('th', {'class':'css-k5ermf ezjvm5n0'}).string] = i.td.get_text()
            except:
                self.info[car_url][i.find('th', {'class':'css-1u7gkch ezjvm5n0'}).string] = i.td.get_text()

        try:
            for i in self.soup.find_all('div',{'class':'css-4hux4x e162wx9x0'}):
                words = i.get_text().split(':')
                for j in range(len(words) - 1):
                    words[1] = words[1].replace('\n', ' ')
                    self.info[car_url][words[0]] = words[1]

            imgs = []
            for i in self.soup.find('div',{'class':'css-1i28qtd e15of25h0'}).find_all('img'):
                imgs.append(i['srcset'].split(' ')[2])
            self.info[car_url]['imgs'] = imgs
        except:
            pass

    def get_all_car_name_and_url(self):
        '''
        Переходит по 100 страницам сайта(максимум, которорый отображается на сайте) и извлекает url автомобилей
        '''

        for i in range(2, 101):
            self.get_car_name_and_url()
            self.next_page(i)

    def save_imgs_in_new_folder(self, car_url):
        '''
        Сохраняет фотографии на каждый атомобиль в отдельную папку.
        '''

        imgs_links = self.info[car_url]['imgs']

        new_folder = os.path.join(self.path_name, 'id_{0}'.format(self.folder_id) + '_' + self.info[car_url]['name'].replace(',', '').replace(' ', '_'))
        if not os.path.exists(new_folder):
            os.makedirs(new_folder)
            self.folder_id += 1

        for number, link in enumerate(imgs_links):
            order_number = number + 1
            img = requests.get(link, stream=True)
            img.raw.decode_content = True
            with open('{0}\\img_{1}.jpg'.format(new_folder, order_number), 'wb') as image_file:
                shutil.copyfileobj(img.raw, image_file)
            del img

    def save_info_in_exel(self):
        '''
        Создает файл exel в рабочей директории и загружает туда всю информацию
        '''

        wb = Workbook()
        exel = wb.active
        exel.title = 'Info'

        row = 1
        col = 1
        for link in self.info:
            for header in self.info[link]:
                if header != 'imgs':
                    exel.cell(row=row, column=col, value=header).alignment = Alignment(horizontal='center')
                    col += 1
            row += 1
            col = 1
            break

        for link in self.info:
            for header in self.info[link]:
                if header != 'imgs':
                    value = self.info[link][header]
                    exel.cell(row=row, column=col, value=value).alignment = Alignment(horizontal='center')
                    col += 1
            row += 1
            col = 1

        for column_cells in exel.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 10
            exel.column_dimensions[column_cells[0].column_letter].width = length

        wb.save('Info.xlsx')

    def finish(self):
        '''
        Закрывает браузер
        '''

        self.driver.quit()

    def execute(self):
        '''
        Выполняет парсинг
        '''

        self.get_all_car_name_and_url()

        for link in self.info:
            self.get_car_info(link)
            try:
                self.save_imgs_in_new_folder(link)
            except:
                pass

        self.save_info_in_exel()

        self.finish()


parser = Parser('https://vladivostok.drom.ru/auto/all/?inomarka=1&unsold=1&distance=500&order=price&tcb=1603558101')
parser.execute()
