from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Border, Side
import json
from bs4 import BeautifulSoup

class Parser:
    '''
    Парсинг сайта 'https://www.finder.fi/search'.
    На вход требует запрос по которому будет произведен поиск на сайте.
    Собирает информацию о людях, которые работают в компаниях (если они есть).
    Информацию сохраняет в json и excel.
    '''
    def __init__(self):
        '''
        Запускает браузер Googl в "безголовом" режиме.
        '''

        self.root_url = 'https://www.finder.fi/search'

        self.info = {}

        google_exe = 'chromedriver.exe'
        options = webdriver.ChromeOptions()
        # full screen
        options.add_argument("--start-maximized")
        options.add_argument('headless')

        self.driver = webdriver.Chrome(google_exe, options=options)

    def search(self, word):
        '''
        Выполняет поиск на сайте по введенному слову от пользователя.
        Переходит в раздел "Компании".
        Получает url страницы, полученной в результате поска.
        '''

        self.word = word

        self.driver.get(self.root_url)

        try:
            text_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#react-toplevel > div > header > div.Header__Content > div > div > input'))
            )
        except:
            return None

        action = ActionChains(self.driver)
        # вводит слово в элемент "text_input"
        action.send_keys_to_element(text_input, word)
        # "нажимает" кнопку Enter
        action.send_keys_to_element(text_input, Keys.RETURN)
        action.perform()

        try:
            companies = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#react-toplevel > div > main > div > div.SearchResultList.SearchResultList--advertisement.SearchResultList--done > div.SearchResultList__TabBar > a:nth-child(2)'))
            )
        except:
            return None

        action = ActionChains(self.driver)
        action.click(companies)
        action.perform()

        # получает url страницы
        self.current_url = self.driver.current_url

    def general_page(self, page_number):
        '''
        Принимает номер страницы и переходит на эту страницу.
        '''

        page_url = self.current_url + '&page=' + str(page_number)
        self.driver.get(page_url)

    def get_company_url(self):
        '''
        Собирает url каждой компании на страничке.
        '''

        soup = BeautifulSoup(self.driver.page_source, 'lxml')
        for company in soup.find_all('a', {'class': 'SearchResult__ProfileLink'}):
            self.info[self.root_url.replace('/search', '') + company['href']] = {}

    def gather_company_url(self):
        '''
        Выясняет какое максимальное количество страниц по данному запросу.
        Переходит по страничкам и собирает url всех компаний.
        '''

        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        pages = []
        for page in soup.find('div', {'class': 'SearchResultList__PageSelection SearchResultList__Navigation__Col'}).find_all('li'):
            try:
                pages.append(int(page.get_text()))
            except:
                pass

        max_page = max(pages)

        for page_number in range(1, max_page + 1):
            try:
                self.general_page(page_number)
                time.sleep(10)
                self.get_company_url()
            except:
                pass

    def get_company_info(self, url):
        '''
        Извлекает информацию о каждой компании и людях, которые там работают и помещает информацию в словарь 'self.info'
        '''

        time.sleep(3)

        self.driver.get(url)

        try:
            show_number = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="react-toplevel"]/div/main/div/div/div[1]/div[1]/div[1]/div/div/div[2]/div[1]/ul/div[1]/li[2]/div[1]/span[2]/span'))
            )
            action = ActionChains(self.driver)
            action.click(show_number)
            action.perform()
        except:
            pass


        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        self.info[url]['company_info'] = {}

        try:
            name = soup.find('div', {'class': 'Profile__Name listing-name'}).get_text()
            self.info[url]['company_info']['Полное имя компании'] = name
        except:
            try:
                name = soup.find('div', {'class': 'Profile__Name Profile__Name--short listing-name'}).get_text()
                self.info[url]['company_info']['Полное имя компании'] = name
            except:
                pass


        try:
            for i in soup.find('ul', {'class': 'fa-ul ContactInformation__List'}).find_all('li'):
                number = i.get_text().split()[-1][:-1]
                break
            self.info[url]['company_info']['Номер'] = number
        except:
            pass


        try:
            phone_number = soup.find('a', {'class': 'listing-phone'})['href'].replace('tel:', '')
            self.info[url]['company_info']['Номер телефона компании'] = phone_number
        except:
            pass


        try:
            address = soup.find('a', {'class': 'SearchResult__Link'}).get_text()
            self.info[url]['company_info']['Адрес'] = address
        except:
            pass

        try:
            site = soup.find('a', {'class': 'listing-website-url'}).get_text()
            self.info[url]['company_info']['Сайт'] = site
        except:
            pass


        try:
            mail = soup.find('a', {'class': 'listing-email'}).get_text()
            self.info[url]['company_info']['email'] = mail
        except:
            pass


        try:
            for tr in soup.find('table', {'class': 'Financials__Table table table-condensed'}).find_all('tr'):

                try:
                    if tr.th.get_text() == 'Liikevaihto (tuhatta euroa)':
                        for td in tr.find_all('td'):
                            income = td.get_text()
                        self.info[url]['company_info']['Оборот'] = income
                except:
                    pass


                try:
                    if tr.th.get_text() == 'Henkilöstö':
                        for td in tr.find_all('td'):
                            employ = td.get_text()
                        self.info[url]['company_info']['Количество рабочих'] = employ
                except:
                    pass
        except:
            pass


        try:
            bosses = []
            for bos in soup.find_all('li', {'class': 'DecisionPerson DecisionPerson--prh'}):
                bosses.append(bos.get_text()[1:])
            self.info[url]['company_info']['Учредители'] = bosses
        except:
            pass


        try:
            persons_url = []
            for person_url in soup.find_all('a', {'class': 'AdditionalContacts__Name'}):
                persons_url.append(self.root_url.replace('/search', '') + person_url['href'])
            self.info[url]['company_info']['persons_url'] = persons_url
        except:
            pass


        time.sleep(0.5)

        try:
            for person_url in self.info[url]['company_info']['persons_url']:

                self.driver.get(person_url)
                self.info[url]['company_info'][person_url] = {}

                time.sleep(1)

                try:
                    show_number = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="react-toplevel"]/div/main/div/div/div[1]/div[1]/div/div/ul/div[1]/li/div[1]/span[2]'))
                    )
                    action = ActionChains(self.driver)
                    action.click(show_number)
                    action.perform()
                except:
                    pass

                soup = BeautifulSoup(self.driver.page_source, 'lxml')

                try:
                    name = soup.find('div', {'class': 'Profile__Name Profile__Name--short listing-name'}).get_text()
                    self.info[url]['company_info'][person_url]['Имя работника'] = name
                except:
                    try:
                        name = soup.find('div', {'class': 'Profile__Name listing-name'}).get_text()
                        self.info[url]['company_info'][person_url]['Имя работника'] = name
                    except:
                        pass

                try:
                    position = soup.find('div', {'class': 'Profile__Title'}).get_text()
                    self.info[url]['company_info'][person_url]['Должность'] = position
                except:
                    pass

                try:
                    phone_number = soup.find('div', {'class': 'PhoneNumber__Unhidden'}).find('a')['href'].replace('tel:', '')
                    self.info[url]['company_info'][person_url]['Номер работника'] = phone_number
                except:
                    pass

                try:
                    email = soup.find('a', {'class': 'listing-email'}).get_text()
                    self.info[url]['company_info'][person_url]['email работника'] = email
                except:
                    pass
        except:
            pass

    def save_in_json(self):
        '''
        Сохраняет self.info в json
        '''

        with open(self.word + '.json', 'w') as file:
            json.dump(self.info, file)

    def save_in_excel(self):
        '''
        Создает файл excel в рабочей директории и загружает туда всю информацию.
        '''

        wb = Workbook()
        excel = wb.active
        excel.title = 'Info'

        light_style = NamedStyle(name='light_style')
        light_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        light_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        light_style.alignment = Alignment(horizontal='center', vertical='center')
        light_style.fill = PatternFill(fgColor="99CCFF", fill_type="solid")
        wb.add_named_style(light_style)

        dark_style = NamedStyle(name='dark_style')
        dark_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        dark_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        dark_style.alignment = Alignment(horizontal='center', vertical='center')
        dark_style.fill = PatternFill(fgColor="33CCCC", fill_type="solid")
        wb.add_named_style(dark_style)

        excel.cell(row=1, column=1, value='Имя человека')
        excel.cell(row=1, column=2, value='Позиция')
        excel.cell(row=1, column=3, value='Телефон')
        excel.cell(row=1, column=4, value='email')
        excel.cell(row=1, column=5, value='Полное имя компании')
        excel.cell(row=1, column=6, value='Номер')
        excel.cell(row=1, column=7, value='Телефон')
        excel.cell(row=1, column=8, value='Адрес')
        excel.cell(row=1, column=9, value='email')
        excel.cell(row=1, column=10, value='Сайт')
        excel.cell(row=1, column=11, value='Оборот')
        excel.cell(row=1, column=12, value='Количество рабочих')
        excel.cell(row=1, column=13, value='Учредители')

        row = 1
        for company_url in self.info:

            try:
                for person_url in self.info[company_url]['company_info']['persons_url']:

                    row += 1

                    try:
                        value = self.info[company_url]['company_info'][person_url]['Имя работника']
                        excel.cell(row=row, column=1, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info'][person_url]['Должность']
                        excel.cell(row=row, column=2, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info'][person_url]['Номер работника']
                        excel.cell(row=row, column=3, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info'][person_url]['email работника']
                        excel.cell(row=row, column=4, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Полное имя компании']
                        excel.cell(row=row, column=5, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Номер']
                        excel.cell(row=row, column=6, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Номер телефона компании']
                        excel.cell(row=row, column=7, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Адрес']
                        excel.cell(row=row, column=8, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['email']
                        excel.cell(row=row, column=9, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Сайт']
                        excel.cell(row=row, column=10, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Оборот']
                        excel.cell(row=row, column=11, value=value)
                    except:
                        pass

                    try:
                        value = self.info[company_url]['company_info']['Количество рабочих']
                        excel.cell(row=row, column=12, value=value)
                    except:
                        pass

                    try:
                        value = '\n'.join(self.info[company_url]['company_info']['Учредители'])
                        excel.cell(row=row, column=13, value=value)
                    except:
                        pass
            except:
                pass

        for column_cells in excel.columns:
            try:
                length = max(len(str(cell.value)) for cell in column_cells) + 16
                excel.column_dimensions[column_cells[0].column_letter].width = length
            except:
                pass

        count = 1
        for row in excel.rows:

            if count % 2 == 0:
                count_cell = 0

                for cell in row:
                    count_cell += 1

                    cell.style = light_style

                    if count_cell == 13:
                        try:
                            length = max(map(len, str(cell.value).splitlines())) + 12
                            excel.column_dimensions[cell.column_letter].width = length
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                        except:
                            pass

            else:
                count_cell = 0

                for cell in row:
                    count_cell += 1

                    cell.style = dark_style

                    if count_cell == 13:
                        try:
                            length = max(map(len, str(cell.value).splitlines())) + 12
                            excel.column_dimensions[cell.column_letter].width = length
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                        except:
                            pass
            count += 1

        wb.save(self.word + '.xlsx')

    def finish(self):
        '''
        Закрывает браузер.
        '''

        self.driver.quit()

    def execut(self):
        '''
        Выполняет парсинг.
        '''

        word = input('Введите запрос:')

        try:
            self.search(word)

            time.sleep(5)

            self.gather_company_url()

            for company_url in self.info:
                try:
                    self.get_company_info(company_url)
                except:
                    pass

            self.save_in_json()
            self.save_in_excel()
        except:
            pass


parser = Parser()
parser.execut()
