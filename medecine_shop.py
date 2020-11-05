from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Border, Side
import json
from bs4 import BeautifulSoup
import os
import requests
import shutil

class Parser:
    '''
    Парсинг сайта 'https://diamarka.com/'. Магазин аптечных товаров.
    Собирает информацию о товарах (Раздел, в которой находится товар, название товара, новая цена, старая цена, описание, фото, ссылка).
    Создает папку в рабочей директории 'imgs', в которую будут загружены фото товаров.
    Информацию сохраняет в json и excel.
    '''

    def __init__(self):
        '''
        Запускает браузер Google в "безголовом" режиме.
        Создает папку 'imgs' в рабочей директории.
        '''

        self.info = {}

        google_exe = 'chromedriver.exe'
        options = webdriver.ChromeOptions()
        # full screen
        options.add_argument("--start-maximized")
        #options.add_argument("headless")

        self.root_url = 'https://diamarka.com'

        self.driver = webdriver.Chrome(google_exe, options=options)
        self.driver.get(self.root_url)

        time.sleep(5)

        current_path = os.getcwd()

        self.imgs_folder = 'imgs'
        path = current_path + '\{}'.format(self.imgs_folder)
        if not os.path.exists(path):
            os.makedirs(path)

    def get_categories_urls(self):
        '''
        Собирает url каждой категории.
        '''

        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        for catalog in soup.find_all('div', {'class': 'catalog-menu__item1'}):
            catalog_name = catalog.a.get_text()
            self.info[catalog_name] = {}

            for category in catalog.find_all('div', {'class': 'catalog-menu__box2'}):
                category_name = category.a.get_text()
                category_url = self.root_url + category.a['href']
                self.info[catalog_name][category_name] = {'url': category_url}

    def get_product_urls(self, category_url, catalog_name, category_name):
        '''
        Собирает url каждого товара со странички.
        Пример аргументов на вход: ("https://diamarka.com/glyukometry/", "Диабетические товары", "Глюкометры ")
        '''

        products_url = []

        self.driver.get(category_url)
        time.sleep(10)
        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        def get_urls(self):
            time.sleep(3)

            soup = BeautifulSoup(self.driver.page_source, 'lxml')

            for product in soup.find_all('div', {'class': 'catalog_item_wrapp item'}):
                product_url = self.root_url + product.a['href']
                products_url.append(product_url)

        get_urls(self)

        try:
            '''
            Если продуктов одной категории больше чем одна страничка, тогда перелистывает их все.
            '''
            max_page = 0
            for page_number in soup.find('div', {'class': 'nums'}).find_all('a', {'class': 'dark_link'}):
                max_page = int(page_number.get_text())

            for page_number in range(2, max_page + 1):

                page_url = category_url + '?PAGEN_2={}'.format(page_number)
                self.driver.get(page_url)
                get_urls(self)
        except:
            pass

        self.info[catalog_name][category_name]['product_urls'] = products_url

    def get_product_info(self, product_url, catalog_name, category_name, product_number):
        '''
        Извлекает информацию о товаре и помещает информацию в словарь 'self.info'
        '''

        self.driver.get(product_url)
        time.sleep(3)
        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        name = soup.find('h1', {'id': 'pagetitle'}).get_text()
        self.info[catalog_name][category_name][str(product_number)] = {}

        self.info[catalog_name][category_name][str(product_number)]['url'] = product_url

        try:
            self.info[catalog_name][category_name][str(product_number)]['Название товара'] = name
        except:
            pass

        count = 0
        for i in soup.find_all('div', {'class': 'price_group'}):
            count += 1

            if count == 1:
                try:
                    new_price = i.find('span', {'class': 'price_value'}).get_text() + i.find('span', {'class': 'price_currency'}).get_text()
                    self.info[catalog_name][category_name][str(product_number)]['Новая цена'] = new_price
                except:
                    pass

            if count == 2:
                try:
                    old_price = i.find('span', {'class': 'price_value'}).get_text() + i.find('span', {'class': 'price_currency'}).get_text()
                    self.info[catalog_name][category_name][str(product_number)]['Старая цена'] = old_price
                except:
                    pass

        try:
            article = soup.find('div', {'class': 'preview_text dotdot'}).get_text()
            self.info[catalog_name][category_name][str(product_number)]['Описание'] = article
        except:
            pass

        try:
            img_url = self.root_url + soup.find('div', {'class': 'slides'}).img['src']
            self.info[catalog_name][category_name][str(product_number)]['Фото'] = '{}.jpg'.format(product_number)

            img = requests.get(img_url, stream=True)
            img.raw.decode_content = True

            with open(r'{}\{}.jpg'.format(self.imgs_folder, product_number), 'wb') as image_file:
                shutil.copyfileobj(img.raw, image_file)
        except :
            pass

    def save_in_json(self):
        '''
        Сохраняет self.info в json
        '''

        with open('Info.json', 'w') as file:
            json.dump(self.info, file)

    def save_in_excel(self):
        '''
        Создает файл excel в рабочей директории и загружает туда всю информацию.
        '''

        wb = Workbook()
        excel = wb.active
        excel.title = 'Info'

        light_style = NamedStyle(name='light_style')
        light_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        light_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        light_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        light_style.fill = PatternFill(fgColor="99CCFF", fill_type="solid")
        wb.add_named_style(light_style)

        dark_style = NamedStyle(name='dark_style')
        dark_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        dark_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        dark_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        dark_style.fill = PatternFill(fgColor="33CCCC", fill_type="solid")
        wb.add_named_style(dark_style)

        header_style = NamedStyle(name='header_style')
        header_style.font = Font(size=20, bold=True)
        bd = Side(style='thick', color='000000')
        header_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        header_style.fill = PatternFill(fgColor="1896a5", fill_type="solid")
        wb.add_named_style(header_style)

        excel.cell(row=1, column=1, value='Раздел 1')
        excel.cell(row=1, column=2, value='Раздел 2')
        excel.cell(row=1, column=3, value='Название товара')
        excel.cell(row=1, column=4, value='Новая цена')
        excel.cell(row=1, column=5, value='Старая цена')
        excel.cell(row=1, column=6, value='Описание')
        excel.cell(row=1, column=7, value='Фото')
        excel.cell(row=1, column=8, value='Ссылка')

        row = 2
        for catalog_name in self.info:
            for category_name in self.info[catalog_name]:
                for product_number in self.info[catalog_name][category_name]:

                    value = catalog_name
                    excel.cell(row=row, column=1, value=value)

                    value = category_name
                    excel.cell(row=row, column=2, value=value)

                    try:
                        value = self.info[catalog_name][category_name][product_number]['Название товара']
                        excel.cell(row=row, column=3, value=value)
                    except:
                        continue

                    try:
                        value = self.info[catalog_name][category_name][product_number]['Новая цена']
                        excel.cell(row=row, column=4, value=value)
                    except:
                        pass

                    try:
                        value = self.info[catalog_name][category_name][product_number]['Старая цена']
                        excel.cell(row=row, column=5, value=value)
                    except:
                        pass

                    try:
                        value = self.info[catalog_name][category_name][product_number]['Описание'].strip()
                        excel.cell(row=row, column=6, value=value)
                    except:
                        pass

                    try:
                        value = self.info[catalog_name][category_name][product_number]['Фото']
                        excel.cell(row=row, column=7, value=value)
                    except:
                        pass

                    try:
                        value = self.info[catalog_name][category_name][product_number]['url']
                        excel.cell(row=row, column=8, value=value)
                    except:
                        pass

                    row += 1

        for column_cells in excel.columns:
            try:
                length = max(max(map(len, str(cell.value).splitlines())) for cell in column_cells) + 8
                excel.column_dimensions[column_cells[0].column_letter].width = length
            except:
                pass

        count = 1
        for row in excel.rows:
            excel.row_dimensions[count].height = 70

            if count % 2 == 0:
                for cell in row:
                    cell.style = light_style

            if count == 1:
                for cell in row:
                    cell.style = header_style

            elif count % 2 != 0:
                for cell in row:
                    cell.style = dark_style

            count += 1

        # замораживает первою строку
        excel.freeze_panes = 'A2'

        wb.save('Info.xlsx')

    def finish(self):
        '''
        Закрывает браузер.
        '''

        self.driver.quit()

    def execute(self):
        '''
        Выполняет парсинг.
        '''

        self.get_categories_urls()

        for catalog_name in self.info:
            for category_name in self.info[catalog_name]:

                category_url = self.info[catalog_name][category_name]['url']
                self.get_product_urls(category_url, catalog_name, category_name)

        count = 1
        for catalog_name in self.info:
            for category_name in self.info[catalog_name]:
                for product_url in self.info[catalog_name][category_name]['product_urls']:

                    try:
                        self.get_product_info(product_url, catalog_name, category_name, count)
                    except:
                        pass

                    count += 1

        self.save_in_json()
        self.save_in_excel()
        self.finish()


parser = Parser()
parser.execute()
