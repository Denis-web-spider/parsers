from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import json
from bs4 import BeautifulSoup, NavigableString, Tag


class Parse:
    '''
    Парсер для сайта https://www.finder.fi/search?what=Rakennusmateriaalit.
    Извлекает всю информацию про каждую компанию и помещает всю информацию в таблицу exel, котороя создается в рабочей директории "Info.xlsx".
    '''

    def __init__(self):
        '''
        Запускает браузер Googl в "безголовом" режиме.
        '''

        self.info = {}

        # используется для ключей в self.info обозначает номер компании на сайте
        self.company_count = 0

        google_exe = 'chromedriver.exe'
        options = webdriver.ChromeOptions()
        # full screen
        options.add_argument("--start-maximized")
        options.add_argument('headless')

        self.driver = webdriver.Chrome(google_exe, options=options)

    def page(self, number_of_page):
        '''
        Принимает в качестве агрумента номер страницы на которую нужно перейти.
        Модифицирует url и открывает страницу в браузере.
        '''

        url = 'https://www.finder.fi/search?what=Rakennusmateriaalit&page={}'.format(number_of_page)
        self.driver.get(url)

    def show_numbers(self):
        '''
        Нажимает на кнопку для отображения полного номера телефона.
        Для этого кнопка должна находится на экране (иначе невозможно нажать).
        '''

        number_order = 1

        while True:
            '''
            Передвигает экран к следуующе кнопке. Кнопка всегда находится на 300px ниже края экрана.
            Если номера и соответсвенно кнопки нету, то нажимается следующая кнопка.
            Кнопок на страничке максимум 25.
            После нажатия всех кнопок метод завершает свою работу.
            '''
            try:

                if number_order > 26:
                    break

                time.sleep(1)

                number_order += 1

                try:
                    button = WebDriverWait(self.driver, 1).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '#react-toplevel > div > main > div > div.SearchResultList.SearchResultList--advertisement.SearchResultList--done > div:nth-child({0}) > div.SearchResult__AddrCol > ul > li.SearchResult__Phone > div > span.PhoneNumber__ShowButton'.format(number_order)))
                        )
                except:
                    continue

                loc = button.location['y'] - 300

                self.driver.execute_script("window.scrollTo(0, {0})".format(loc))

                time.sleep(1)

                action = ActionChains(self.driver)
                action.click(button)
                action.perform()
            except:
                pass

    def get_info(self):
        '''
        Извлекает всю информацию про фирму и помещает её в "self.info['Номер компании']['Характеристика'] = Значение".
        '''

        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        for i in soup.find_all('div', {'class': 'SearchResult SearchResult--distinctive SearchResult--compact'}):

            self.company_count += 1

            self.info[str(self.company_count)] = {}

            self.info[str(self.company_count)]['Компания'] = i.find('div', {'class': 'SearchResult__Name'}).get_text()

            for k in i.find_all('div', {'class': 'text-muted'}):
                if isinstance(k, NavigableString):
                    continue
                if isinstance(k, Tag):
                    if not k.get_text().split(' ')[-1].isalpha():
                        self.info[str(self.company_count)]['Номер'] = k.get_text().split(' ')[-1]

            self.info[str(self.company_count)]['Адресс'] = i.find('a', {'class': 'SearchResult__Link'}).get_text()

            for j in i.find_all('a', {'class': 'undefined'}):
                if j['href'][0:3] == 'tel':
                    self.info[str(self.company_count)]['Телефон'] = j['href'].replace('tel:', '')
                else:
                    self.info[str(self.company_count)]['Сайт'] = j['href']

        for i in soup.find_all('div', {'class': 'SearchResult SearchResult--compact'}):

            self.company_count += 1

            self.info[str(self.company_count)] = {}

            self.info[str(self.company_count)]['Компания'] = i.find('div', {'class': 'SearchResult__Name'}).get_text()

            for k in i.find_all('div', {'class': 'text-muted'}):
                if isinstance(k, NavigableString):
                    continue
                if isinstance(k, Tag):
                    if not k.get_text().split(' ')[-1].isalpha():
                        self.info[str(self.company_count)]['Номер'] = k.get_text().split(' ')[-1]

            try:
                self.info[str(self.company_count)]['Адресс'] = i.find('a', {'class': 'SearchResult__Link'}).get_text()
            except:
                pass

            try:
                for j in i.find_all('a', {'class': 'undefined'}):

                    if j['href'][0:3] == 'tel':
                        try:
                            self.info[str(self.company_count)]['Телефон'] = j['href'].replace('tel:', '')
                        except:
                            pass

                    else:
                        try:
                            self.info[str(self.company_count)]['Сайт'] = j['href']
                        except:
                            pass
            except:
                pass

    def save_in_excel(self):
        '''
        Создает файл exel в рабочей директории и загружает туда всю информацию.
        '''

        wb = Workbook()
        exel = wb.active
        exel.title = 'Info'

        row = 1
        col = 1
        for company_number in self.info:
            for header in self.info[company_number]:
                exel.cell(row=row, column=col, value=header).alignment = Alignment(horizontal='center')
                exel.cell(row=row, column=col).font = Font(size='14')
                col += 1
            row += 1
            break

        for company_number in self.info:

            try:
                value = self.info[company_number]['Компания']
                exel.cell(row=row, column=1, value=value).alignment = Alignment(horizontal='center', vertical='center')
                exel.cell(row=row, column=1).font = Font(size='14')
            except:
                pass

            try:
                if not self.info[company_number]['Номер'][0].isdigit():
                    print(self.info[company_number]['Номер'])
                    value = ''
                else:
                    value = self.info[company_number]['Номер']
                exel.cell(row=row, column=2, value=value).alignment = Alignment(horizontal='center', vertical='center')
                exel.cell(row=row, column=2).font = Font(size='14')
            except:
                pass

            try:
                value = self.info[company_number]['Адресс']
                exel.cell(row=row, column=3, value=value).alignment = Alignment(horizontal='center', vertical='center')
                exel.cell(row=row, column=3).font = Font(size='14')
            except:
                pass

            try:
                value = self.info[company_number]['Сайт']
                exel.cell(row=row, column=4, value=value).alignment = Alignment(horizontal='center', vertical='center')
                exel.cell(row=row, column=4).font = Font(size='14')
            except:
                pass

            try:
                value = self.info[company_number]['Телефон']
                exel.cell(row=row, column=5, value=value).alignment = Alignment(horizontal='center', vertical='center')
                exel.cell(row=row, column=5).font = Font(size='14')
            except:
                pass

            row += 1

        for column_cells in exel.columns:
            try:
                length = max(len(str(cell.value)) for cell in column_cells) + 8
                exel.column_dimensions[column_cells[0].column_letter].width = length
            except:
                continue

        count = 1
        for row in exel.rows:

            if count % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(fgColor="99CCFF", fill_type="solid")

            else:
                for cell in row:
                    cell.fill = PatternFill(fgColor="33CCCC", fill_type="solid")

            count += 1

        wb.save('Info.xlsx')

    def finish(self):
        '''
        Закрывает браузер.
        '''

        self.driver.quit()

    def save_in_json(self):
        '''
        Сохраняет информяцию в 'inform.json'.
        '''

        with open('inform.json', 'w') as f:
            json.dump(self.info, f)


    def execute(self, pages):
        '''
        Выполняет парсинг.
        '''

        for page in range(1, pages + 1):

            try:
                self.page(page)
                self.show_numbers()
                self.get_info()
            except:
                print('exception!', page)

        self.save_in_json()
        self.save_in_excel()
        self.finish()


parse = Parse()
parse.execute(221)
