from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import mysql.connector
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Border, Side
from bs4 import BeautifulSoup
import os
import requests
import shutil

class Parser:
    '''
    Парсинг сайта 'https://www.its.co.uk/Power-Tools/'. Магазин строительных инструментов.
    Собирает информацию о товарах (id, название товара, цена, ссылка на фото).
    Создает папку в рабочей директории 'Главная', в которую будут созданы папки с названиями категорий, в которые будут загружены фото товаров.
    Информацию сохраняет в mysql и excel.
    '''

    def __init__(self, db_name='power_tools', host='localhost', user='root', password='my_password', table_name='Products'):
        '''
        Запускает браузер Google в "безголовом" режиме.
        Создает папку 'Главная' в рабочей директории.
        '''

        self.database = db_name
        self.host = host
        self.user = user
        self.password = password
        self.table_name = table_name

        google_exe = 'chromedriver.exe'
        options = webdriver.ChromeOptions()
        # full screen
        options.add_argument("--start-maximized")
        options.add_argument("headless")

        self.root_url = 'https://www.its.co.uk'
        self.main_url = 'https://www.its.co.uk/Power-Tools/'

        self.category_name = {}

        self.driver = webdriver.Chrome(google_exe, options=options)

        self.driver.get(self.main_url)

        time.sleep(5)

        current_path = os.getcwd()
        self.imgs_folder = 'Главная'
        self.path = current_path + '\{}'.format(self.imgs_folder)
        if not os.path.exists(self.path):
            os.makedirs(self.path)

    def create_db(self):
        '''
        Создает базу данных mysql.
        '''

        db = mysql.connector.connect(
            host=self.host,
            user=self.user,
            password=self.password
        )

        cursor = db.cursor()

        try:
            cursor.execute('CREATE DATABASE {}'.format(self.database))
        except:
            print('База данных {} уже существует'.format(self.database))

    def connect_to_database(self):
        '''
        Соединяется с базой данных.
        '''

        self.db = mysql.connector.connect(
            host=self.host,
            user=self.user,
            password=self.password,
            database=self.database
        )

        self.cursor = self.db.cursor()

    def create_table_in_db(self):
        '''
        Создает таблицу в базе данных. Если таблица существует, то перезаписывает её.
        '''

        self.cursor.execute('DROP TABLE IF EXISTS {}'.format(self.table_name))

        self.cursor.execute('''CREATE TABLE {} (
                                                id VARCHAR(255),
                                                Название_товара VARCHAR(255),
                                                Цена VARCHAR(255),
                                                Ссылки_на_фото_товаров VARCHAR(255)
        )'''.format(self.table_name))

        #self.cursor.execute('DESCRIBE {}'.format(self.table_name))

        #for i in self.cursor:
            #print(i)

    def show_info_in_db_table(self):
        '''
        Извлекает всю информацию из базы данных и возвращает курсор.
        '''

        self.cursor.execute('SELECT * FROM {}'.format(self.table_name))

        #for i in self.cursor:
            #print(i)

        return self.cursor

    def insert_info_in_db_table(self, id='NULL', name='NULL', price='NULL', href='NULL'):
        '''
        Вставляет строку в таблицу.
        '''

        sql = 'INSERT INTO {} VALUES (%s, %s, %s, %s)'.format(self.table_name)
        val = (id, name, price, href)
        self.cursor.execute(sql, val)

        self.db.commit()

    def get_categories_url_and_name(self):
        '''
        Извлекает url и название каждой категории и помезает в "self.category_name[url] = category_name"
        '''

        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        for i in soup.find('div', {'class': 'M1 sublvl tag'}).find_all('a'):
            url = self.root_url + i['href']
            category_name = i.get_text()
            self.category_name[url] = category_name

    def get_product_info(self, category_url):
        '''
        Извлекает информацию о каждом продукте в категории. Помещает информацию в базу данных.
        Создает папку с названием категории в папку "Главная", помещает туда фотографии каждого товара.
        '''

        imgs_folder = self.category_name[category_url]
        path = self.path + '\{}'.format(imgs_folder)
        if not os.path.exists(path):
            os.makedirs(path)

        self.driver.get(category_url)

        time.sleep(2)

        try:
            show_all_butomn = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#ctl00_cph1_plpGrid_Paging1_PagingLink14'))
            )

            loc = show_all_butomn.location['y'] - 300

            self.driver.execute_script("window.scrollTo(0, {0})".format(loc))

            time.sleep(2)

            action = ActionChains(self.driver)
            action.move_to_element(show_all_butomn)
            action.click()
            action.perform()
        except:
            pass

        time.sleep(10)

        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        for i in soup.find('div', {'class': 'LISTCONTAINER'}).find_all('li'):

            time.sleep(1)

            try:
                not_available = i.find('span', {'class': 'No Longer Available'})

                if not_available != None:
                    continue
            except:
                pass

            try:
                name = '' + i.find('span', {'class': 'WebBrand'}).get_text() + '\n' + i.find('span', {'class': 'desc h2'}).get_text()
            except:
                name = 'NULL'

            try:
                id = i.find('span', {'class': 'CatRefGridList'}).get_text().replace('(', '').replace(')', '').strip()
            except:
                id = 'NULL'

            try:
                name = name.replace('({})'.format(id), '')
            except:
                name = 'NULL'

            try:
                price = i.find('span', {'class': 'VatEx'}).get_text()[1:]
            except:
                price = 'NULL'

            try:
                img_href = i.find('img', {'class': 'imgMain'})['src']

                img = requests.get(img_href, stream=True)
                img.raw.decode_content = True

                with open(r'{}\{}.jpg'.format(self.imgs_folder + '\\' + imgs_folder, name.splitlines()[0].replace('/', '_')), 'wb') as image_file:
                    shutil.copyfileobj(img.raw, image_file)
            except:
                pass

            self.insert_info_in_db_table(id, name, price, img_href)

    def save_info_in_excel(self):
        '''
        Создает файл excel в рабочей директории и загружает туда всю информацию.
        '''

        cursor = self.show_info_in_db_table()

        wb = Workbook()
        excel = wb.active
        excel.title = 'Info'

        light_style = NamedStyle(name='light_style')
        light_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        light_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        light_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        light_style.fill = PatternFill(fgColor="99CCFF", fill_type="solid")
        wb.add_named_style(light_style)

        dark_style = NamedStyle(name='dark_style')
        dark_style.font = Font(size=16)
        bd = Side(style='thick', color='000000')
        dark_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        dark_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        dark_style.fill = PatternFill(fgColor="33CCCC", fill_type="solid")
        wb.add_named_style(dark_style)

        header_style = NamedStyle(name='header_style')
        header_style.font = Font(size=20, bold=True)
        bd = Side(style='thick', color='000000')
        header_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        header_style.fill = PatternFill(fgColor="1896a5", fill_type="solid")
        wb.add_named_style(header_style)

        excel.cell(row=1, column=1, value='id')
        excel.cell(row=1, column=2, value='Название товара')
        excel.cell(row=1, column=3, value='Цена')
        excel.cell(row=1, column=4, value='Ссылка на фото товара')

        row = 2
        for row_info in cursor:

            id = row_info[0]
            name = row_info[1]
            price = row_info[2]
            href = row_info[3]

            excel.cell(row=row, column=1, value=id)
            excel.cell(row=row, column=2, value=name)
            excel.cell(row=row, column=3, value=price)
            excel.cell(row=row, column=4, value=href)

            row += 1

        count = 0
        for column_cells in excel.columns:

            count += 1

            if count == 4:
                try:
                    length = 120
                    excel.column_dimensions[column_cells[0].column_letter].width = length
                except:
                    pass

            else:
                try:
                    length = max(max(map(len, str(cell.value).splitlines())) for cell in column_cells) + 8
                    excel.column_dimensions[column_cells[0].column_letter].width = length
                except:
                    pass

        count = 1
        for row in excel.rows:

            column_count = 0

            excel.row_dimensions[count].height = 70

            if count % 2 == 0:
                for cell in row:
                    column_count += 1

                    cell.style = light_style

                    if column_count == 4:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

            if count == 1:
                for cell in row:
                    cell.style = header_style

            elif count % 2 != 0:
                for cell in row:
                    column_count += 1

                    cell.style = dark_style

                    if column_count == 4:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

            count += 1

        # замораживает первою строку
        excel.freeze_panes = 'A2'

        wb.save('Info.xlsx')

    def finish(self):
        '''
        Закрывает браузер.
        '''

        self.driver.quit()

    def execute(self):
        '''
        Выполняет парсинг.
        '''

        self.connect_to_database()
        self.create_table_in_db()

        try:
            self.get_categories_url_and_name()
        except:
            pass

        for url in self.category_name:

            try:
                self.get_product_info(url)
            except:
                pass

        self.save_info_in_excel()
        self.finish()


parser = Parser()
parser.execute()
